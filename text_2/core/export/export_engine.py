# -*- coding: utf-8 -*-
"""
导出引擎：将对比/校验结果输出为 PDF（带格式）、Excel（含标记与高亮）、CSV（纯数据）。
差异行用颜色区分：新增、删除、修改；校验结果高亮出错字段并标注规则名。
"""

from typing import List, Optional, Dict, Any, Callable
import os
import pandas as pd

try:
    from config import DIFF_ADDED, DIFF_DELETED, DIFF_MODIFIED, EXPORT_DIR
except ImportError:
    DIFF_ADDED = "added"
    DIFF_DELETED = "deleted"
    DIFF_MODIFIED = "modified"
    EXPORT_DIR = "user_data/exports"


# -----------------------------------------------------------------------------
# 颜色常量（新增/删除/修改）
# -----------------------------------------------------------------------------
COLOR_ADDED = "FF00FF00"   # 浅绿
COLOR_DELETED = "FFFF0000" # 浅红
COLOR_MODIFIED = "FFFFFF00" # 浅黄
COLOR_ERROR_FIELD = "FFFF9999"  # 出错字段高亮


class ExportError(Exception):
    """导出失败时抛出。"""
    pass


def _ensure_export_path(path: str) -> str:
    """若 path 仅为文件名，则放到 EXPORT_DIR。"""
    if not os.path.dirname(path):
        path = os.path.join(EXPORT_DIR, path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    return path


def _export_excel_xlsxwriter(
    df: pd.DataFrame,
    path: str,
    cols: list,
    header: list,
    diff_type_column: str,
    changed_fields_column: str,
    violations_by_row: Optional[Dict[int, List[Any]]],
    progress_callback: Optional[Callable[[int, str], None]],
) -> None:
    """使用 xlsxwriter 批量写入，按行设置格式，显著加快大表导出。"""
    import xlsxwriter
    wb = xlsxwriter.Workbook(path, {"constant_memory": True})
    ws = wb.add_worksheet("对比结果")
    fmt_added = wb.add_format({"bg_color": "#90EE90"})
    fmt_deleted = wb.add_format({"bg_color": "#FFB6C1"})
    fmt_modified = wb.add_format({"bg_color": "#FFFFE0"})
    fmt_error = wb.add_format({"bg_color": "#FF9999"})

    for col, h in enumerate(header):
        ws.write_string(0, col, str(h))
    n = len(df)
    batch = 500
    for idx in range(n):
        if progress_callback and idx > 0 and idx % batch == 0:
            progress_callback(int(90 * idx / max(n, 1)), "写入行…")
        row = df.iloc[idx]
        diff_type = row.get(diff_type_column, "")
        out_row = [row.get(c, "") for c in cols] + [diff_type]
        if changed_fields_column in df.columns:
            cf = row.get(changed_fields_column, [])
            out_row.append(str(cf) if not isinstance(cf, str) else cf)
        for col, v in enumerate(out_row):
            ws.write(idx + 1, col, v)
        row_num = idx + 1
        if diff_type == DIFF_ADDED:
            ws.set_row(row_num, None, fmt_added)
        elif diff_type == DIFF_DELETED:
            ws.set_row(row_num, None, fmt_deleted)
        elif diff_type == DIFF_MODIFIED:
            ws.set_row(row_num, None, fmt_modified)
        elif violations_by_row and int(df.index[idx]) in violations_by_row:
            ws.set_row(row_num, None, fmt_error)
    if progress_callback:
        progress_callback(95, "保存文件…")
    wb.close()


def export_to_excel(
    df: pd.DataFrame,
    path: str,
    diff_type_column: str = "__diff_type__",
    changed_fields_column: str = "__changed_fields__",
    violations_by_row: Optional[Dict[int, List[Any]]] = None,
    progress_callback: Optional[Callable[[int, str], None]] = None,
) -> str:
    """
    导出为 Excel，带差异高亮。优先使用 xlsxwriter（更快），否则回退 openpyxl。
    """
    path = _ensure_export_path(path)
    if not path.lower().endswith(".xlsx"):
        path = path + ".xlsx" if not path.endswith(".xlsx") else path

    cols = [c for c in df.columns if c != diff_type_column and c != changed_fields_column]
    header = cols + [diff_type_column]
    if changed_fields_column in df.columns:
        header = cols + [diff_type_column, changed_fields_column]

    try:
        import xlsxwriter
        _export_excel_xlsxwriter(
            df, path, cols, header, diff_type_column, changed_fields_column,
            violations_by_row, progress_callback,
        )
        return path
    except ImportError:
        pass

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        raise ExportError("请安装 openpyxl 或 xlsxwriter: pip install openpyxl xlsxwriter")

    wb = Workbook()
    ws = wb.active
    ws.title = "对比结果"
    fill_added = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fill_deleted = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    fill_modified = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    fill_error = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    ws.append(header)
    n = len(df)
    for idx in range(n):
        if progress_callback and idx % 500 == 0 and idx > 0:
            progress_callback(int(90 * idx / max(n, 1)), "写入行…")
        row = df.iloc[idx]
        diff_type = row.get(diff_type_column, "")
        changed_fields = row.get(changed_fields_column, [])
        if isinstance(changed_fields, str):
            try:
                import ast
                changed_fields = ast.literal_eval(changed_fields) if changed_fields else []
            except Exception:
                changed_fields = []
        row_fill = None
        if diff_type == DIFF_ADDED:
            row_fill = fill_added
        elif diff_type == DIFF_DELETED:
            row_fill = fill_deleted
        elif diff_type == DIFF_MODIFIED:
            row_fill = fill_modified

        out_row = [row.get(c, "") for c in cols] + [diff_type]
        if changed_fields_column in df.columns:
            out_row.append(str(changed_fields))
        ws.append(out_row)
        row_num = ws.max_row
        if row_fill or (violations_by_row and int(df.index[idx]) in violations_by_row):
            for col_idx, c in enumerate(cols, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                if row_fill:
                    cell.fill = row_fill
                if violations_by_row:
                    viols = violations_by_row.get(int(df.index[idx]), [])
                    for v in viols:
                        if hasattr(v, "error_fields") and c in getattr(v, "error_fields", []):
                            cell.fill = fill_error
                            break
    if progress_callback:
        progress_callback(95, "保存文件…")
    try:
        wb.save(path)
    except Exception as e:
        raise ExportError(f"保存 Excel 失败: {e}") from e
    return path


def export_to_csv(
    df: pd.DataFrame,
    path: str,
    diff_type_column: str = "__diff_type__",
    encoding: str = "utf-8-sig",
) -> str:
    """导出为 CSV，含 __diff_type__ 列，无颜色。"""
    path = _ensure_export_path(path)
    if not path.lower().endswith(".csv"):
        path = path + ".csv" if not path.endswith(".csv") else path
    try:
        df.to_csv(path, index=False, encoding=encoding)
    except Exception as e:
        raise ExportError(f"保存 CSV 失败: {e}") from e
    return path


def export_to_pdf(
    df: pd.DataFrame,
    path: str,
    title: str = "清单对比结果",
    diff_type_column: str = "__diff_type__",
    changed_fields_column: str = "__changed_fields__",
    violations_by_row: Optional[Dict[int, List[Any]]] = None,
) -> str:
    """
    导出为 PDF，带简单表格格式与差异说明（颜色用文字标注：新增/删除/修改）。
    """
    path = _ensure_export_path(path)
    if not path.lower().endswith(".pdf"):
        path = path + ".pdf"

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
    except ImportError:
        raise ExportError("请安装 reportlab: pip install reportlab")

    doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    cols = [c for c in df.columns if c not in (diff_type_column, changed_fields_column)]
    header = cols + ["差异类型"]
    data = [header]
    for i, row in df.iterrows():
        diff_type = row.get(diff_type_column, "")
        diff_label = {"added": "新增", "deleted": "删除", "modified": "修改", "unchanged": "未变"}.get(diff_type, diff_type)
        data.append([str(row.get(c, ""))[:20] for c in cols] + [diff_label])

    style_list = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]
    for r in range(1, len(data)):
        diff_val = data[r][-1]
        if diff_val == "新增":
            style_list.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#90EE90")))
        elif diff_val == "删除":
            style_list.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FFB6C1")))
        elif diff_val == "修改":
            style_list.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FFFFE0")))
        else:
            style_list.append(("ROWBACKGROUNDS", (0, r), (-1, r), [colors.white]))
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(style_list))
    story.append(t)
    doc.build(story)

    return path


def export_diff_result(
    diff_result: Any,
    path: str,
    format: str = "xlsx",
    violations_by_row: Optional[Dict[int, List[Any]]] = None,
) -> str:
    """
    将 DiffResult 导出为指定格式。
    format: xlsx | csv | pdf
    """
    df = diff_result.to_dataframe() if hasattr(diff_result, "to_dataframe") else diff_result
    if format.lower() == "csv":
        return export_to_csv(df, path)
    if format.lower() == "pdf":
        return export_to_pdf(df, path, violations_by_row=violations_by_row)
    return export_to_excel(df, path, violations_by_row=violations_by_row)
