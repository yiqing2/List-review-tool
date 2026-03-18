# -*- coding: utf-8 -*-
"""
多格式表格解析模块
Multi-format table parser: Excel, CSV, TSV, Word.
支持复杂多级表头：自动检测表头行数，生成层级列名（如 "系统|专业"）。
"""

import os
import re
from typing import List, Tuple, Optional, Any
import pandas as pd

class ParserError(Exception):
    """解析过程发生的错误（文件损坏、格式不支持等）。"""
    pass


def _normalize_ext(path: str) -> str:
    """获取并统一为小写扩展名。"""
    return os.path.splitext(path)[1].lower()


def _detect_header_rows_excel(df_raw: pd.DataFrame, max_rows: int = 10) -> int:
    """
    启发式检测 Excel 表头行数（多级表头）。
    规则：前几行中，某行之后首次出现“多数列为非空且类型一致”则视为数据起始行，
    该行之上均为表头。若无法判定则返回 0（单行表头）。
    """
    if df_raw.empty or len(df_raw) < 2:
        return 0
    check = df_raw.head(max_rows)
    num_re = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")

    def _is_numeric_like(x: Any) -> bool:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return False
        s = str(x).strip()
        if s == "":
            return False
        # 允许 0.00E+00 这类科学计数法
        return bool(num_re.match(s))

    # 优先：找“数据起始行”= 本行大多数非空为数值，且下一行也像数据
    for start in range(1, len(check) - 1):
        row = check.iloc[start]
        row2 = check.iloc[start + 1]
        non_empty = row.notna() & (row.astype(str).str.strip() != "")
        non_empty2 = row2.notna() & (row2.astype(str).str.strip() != "")
        if non_empty.sum() == 0 or non_empty2.sum() == 0:
            continue
        numeric_cnt = sum(_is_numeric_like(row.iloc[i]) for i in range(len(row)) if non_empty.iloc[i])
        numeric_cnt2 = sum(_is_numeric_like(row2.iloc[i]) for i in range(len(row2)) if non_empty2.iloc[i])
        if numeric_cnt >= max(1, non_empty.sum() // 2) and numeric_cnt2 >= max(1, non_empty2.sum() // 2):
            return start

    # 回退：仅按“非空占比”判断
    for start in range(1, len(check)):
        # 从 start 行开始当作数据：检查该行及下一行是否像数据（数值/字符串混合，非全空）
        row = check.iloc[start]
        non_empty = row.notna() & (row.astype(str).str.strip() != "")
        if non_empty.sum() >= max(1, len(check.columns) // 2):
            return start
    return 0


def _flatten_headers(header_rows: pd.DataFrame, fill_empty: str = "") -> List[str]:
    """
    将多行表头合并为单一层级列名，用 | 连接。
    例如：["系统", "专业"] -> "系统|专业"；空单元格向下/向右填充。
    """
    if header_rows.empty:
        return []
    # 向下填充：上一行有值则继承
    filled = header_rows.copy()
    for c in filled.columns:
        prev = ""
        for r in range(len(filled)):
            v = filled.iloc[r, c]
            if pd.isna(v) or str(v).strip() == "":
                filled.iloc[r, c] = prev
            else:
                prev = str(v).strip()
                filled.iloc[r, c] = prev
    # 按行用 | 连接，得到每列的多级名
    parts = []
    for c in filled.columns:
        parts.append("|".join(str(filled.iloc[r, c]).strip() or fill_empty for r in range(len(filled))))
    return parts


def _simplify_flattened_headers(headers: List[str]) -> List[str]:
    """
    对扁平化后的多级列名做“冗余层级”简化，解决如下常见情况：
    - 顶部一行是整表标题（如 title/标题），下方一行才是实际列名：title|CH1 -> CH1
    - 顶部空、下一行才是列名：|TIME -> TIME

    注意：若多级表头确实表达了层级含义（如 系统|专业），则保留，不做简化。
    """
    if not headers:
        return []
    # 统计第一层的分布，用于判断是否“全表统一标题层”
    first_parts = []
    split_parts = []
    first_parts_multilevel = []
    for h in headers:
        raw = str(h or "").strip()
        parts = [p.strip() for p in raw.split("|")]
        parts = [p for p in parts if p]  # 去掉空层级
        split_parts.append(parts)
        first_parts.append(parts[0] if parts else "")
        if len(parts) >= 2 and parts[0]:
            first_parts_multilevel.append(parts[0])
    generic_titles = {"title", "table", "sheet", "标题", "表题", "表名"}

    simplified = []
    for parts in split_parts:
        if not parts:
            simplified.append("")
            continue
        # 单层：不处理
        if len(parts) == 1:
            simplified.append(parts[0])
            continue
        # 多层：若顶层看起来是“整表标题层”（如 title/标题），则去掉
        if parts and parts[0].strip().lower() in generic_titles:
            parts = parts[1:] if len(parts) > 1 else parts
        # 若去掉后仍有多层，保留层级；否则用最后一层
        if len(parts) >= 2:
            simplified.append("|".join(parts))
        else:
            simplified.append(parts[-1])
    return simplified


def _col_letter_to_index(ref: str) -> int:
    """将单元格引用中的列字母转为索引，如 A->0, B->1, AA->26。"""
    col = ""
    for ch in ref:
        if ch.isalpha():
            col += ch
        else:
            break
    i = 0
    for c in col.upper():
        i = i * 26 + (ord(c) - ord("A") + 1)
    return i - 1


def _read_xlsx_xml_direct(path: str, max_rows: int) -> pd.DataFrame:
    """
    直接解析 xlsx 内 sheet 的 XML，完全绕过 openpyxl 的 dimension，确保读出文件中全部行。
    xlsx 实为 zip，内含 xl/worksheets/sheet1.xml 和 xl/sharedStrings.xml。
    """
    import zipfile
    import xml.etree.ElementTree as ET
    path = os.path.abspath(path)
    target = 1 + max_rows
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    try:
        with zipfile.ZipFile(path, "r") as z:
            names = z.namelist()
            sheet_path = None
            for n in names:
                if "xl/worksheets/sheet" in n and n.endswith(".xml") and "chart" not in n:
                    sheet_path = n
                    break
            if not sheet_path:
                raise ParserError("xlsx 中未找到工作表 XML")
            shared_strings = []
            if "xl/sharedStrings.xml" in names:
                with z.open("xl/sharedStrings.xml") as sf:
                    root = ET.parse(sf).getroot()
                    for si in root.findall("{%s}si" % NS):
                        parts = si.findall(".//{%s}t" % NS)
                        shared_strings.append("".join(t.text or "" for t in parts))
            rows = []
            with z.open(sheet_path) as f:
                for event, elem in ET.iterparse(f, events=("end",)):
                    if elem.tag != "{%s}row" % NS:
                        elem.clear()
                        continue
                    # 按列索引组装该行（cell 的 r 为 A1, B1 等，可能不连续）
                    row_dict = {}
                    for c in elem:
                        if c.tag != "{%s}c" % NS:
                            continue
                        r = c.get("r", "")
                        if not r:
                            continue
                        col_idx = _col_letter_to_index(r)
                        cell_type = c.get("t", "")
                        v_elem = c.find("{%s}v" % NS)
                        if v_elem is not None and v_elem.text is not None:
                            if cell_type == "s":
                                idx = int(v_elem.text)
                                row_dict[col_idx] = shared_strings[idx] if idx < len(shared_strings) else ""
                            else:
                                row_dict[col_idx] = v_elem.text.strip()
                        else:
                            row_dict[col_idx] = ""
                    max_col = max(row_dict.keys()) if row_dict else 0
                    row_list = [row_dict.get(j, "") for j in range(max_col + 1)]
                    rows.append(row_list)
                    elem.clear()
                    if len(rows) >= target:
                        break
    except zipfile.BadZipFile as e:
        raise ParserError(f"无法作为 ZIP 打开 xlsx: {e}") from e
    except Exception as e:
        raise ParserError(f"解析 xlsx 内部 XML 失败: {e}") from e
    if not rows:
        return pd.DataFrame()
    n_cols = max(len(r) for r in rows)
    for i in range(len(rows)):
        rows[i] = rows[i] + [""] * (n_cols - len(rows[i]))
    return pd.DataFrame(rows)


def _read_xlsx_stream(path: str, max_rows: int) -> pd.DataFrame:
    """
    用 openpyxl read_only 流式读取 .xlsx，并 reset_dimensions 以突破错误的“已用范围”。
    若仍只读到 2 千多行，会回退到直接解析 sheet XML（_read_xlsx_xml_direct）。
    返回 raw DataFrame（header=None 风格），行数最多 1+max_rows。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ParserError("请安装 openpyxl: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    try:
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        else:
            try:
                object.__setattr__(ws, "max_row", None)
                object.__setattr__(ws, "max_column", None)
            except Exception:
                pass
    except Exception:
        pass
    target = 1 + max_rows
    rows = []
    try:
        # 不传 max_row，让迭代器尽量遍历 XML 中全部行（依赖 reset_dimensions 生效）
        for row in ws.iter_rows(min_row=1, values_only=True):
            row = list(row) if row else []
            rows.append(["" if v is None else str(v).strip() for v in row])
            if len(rows) >= target:
                break
    finally:
        wb.close()
    # 若 openpyxl 仍只返回 2 千多行，改用直接解析 XML
    if rows and len(rows) < target and len(rows) < 5000:
        try:
            direct = _read_xlsx_xml_direct(path, max_rows)
            if not direct.empty and len(direct) > len(rows):
                return direct
        except Exception:
            pass
    if not rows:
        return pd.DataFrame()
    n_cols = max(len(r) for r in rows)
    for i in range(len(rows)):
        rows[i] = rows[i] + [""] * (n_cols - len(rows[i]))
    return pd.DataFrame(rows)


def _read_excel_safe(
    path: str,
    header_rows: Optional[int] = None,
    max_rows: Optional[int] = None,
    skip_top_rows: int = 0,
) -> Tuple[pd.DataFrame, List[str], int]:
    """
    读取 Excel（.xlsx/.xls），返回 (数据 DataFrame, 列名列表, 实际表头行数)。
    若指定 max_rows（仅 .xlsx）：先用 openpyxl 流式+reset_dimensions 尽量读出全部行，
    若仍不足则补空行并加 __row_index__，保证返回至少 max_rows 行且键唯一。
    """
    ext = _normalize_ext(path)
    raw = None
    if ext == ".xlsx" and max_rows is not None and max_rows > 0:
        # 优先直接解析 sheet XML，不依赖 dimension，可读出全部 12 万行
        try:
            raw = _read_xlsx_xml_direct(path, max_rows)
            # 若只读到 0/1 列（XML 解析常漏列），回退到流式或 read_excel，保证数据列完整（图3 形式）
            if raw is not None and not raw.empty and raw.shape[1] < 2:
                raw = None
        except ParserError:
            raise
        except Exception:
            raw = None
        if raw is None or raw.empty:
            try:
                raw = _read_xlsx_stream(path, max_rows)
                if raw is not None and not raw.empty and raw.shape[1] < 2:
                    raw = None
            except ParserError:
                raise
            except Exception:
                raw = None
    if raw is None or raw.empty:
        try:
            if ext == ".xlsx":
                raw = pd.read_excel(path, header=None, engine="openpyxl", dtype=str)
            elif ext == ".xls":
                raw = pd.read_excel(path, header=None, engine="xlrd", dtype=str)
            else:
                raise ParserError(f"不支持的 Excel 扩展: {ext}")
        except Exception as e:
            raise ParserError(f"读取 Excel 失败: {path}, 错误: {e}") from e

    if raw.empty:
        return pd.DataFrame(), [], 0

    st = int(skip_top_rows or 0)
    if st > 0:
        if len(raw) <= st:
            raise ParserError(f"跳过顶部 {st} 行后无剩余数据，请减小「跳过顶部行数」")
        raw = raw.iloc[st:].reset_index(drop=True)

    if header_rows is None:
        header_rows = _detect_header_rows_excel(raw)
    header_rows = min(header_rows, len(raw) - 1) if len(raw) > 1 else 0

    if header_rows == 0:
        headers = [str(c).strip() for c in raw.iloc[0].tolist()]
        data = raw.iloc[1:].reset_index(drop=True)
        data.columns = range(len(headers))
        data_start_row_1based = st + 2  # 原文件行号（1-based）：跳过 st 行 + 1 行表头
    else:
        header_df = raw.iloc[:header_rows]
        headers = _flatten_headers(header_df)
        data = raw.iloc[header_rows:].reset_index(drop=True)
        data.columns = range(len(headers))
        data_start_row_1based = st + header_rows + 1  # 跳过 st 行 + header_rows 行表头

    headers = _simplify_flattened_headers(headers)

    # 列名去重
    seen = {}
    unique_headers = []
    for h in headers:
        h = str(h).strip()
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    data.columns = unique_headers

    # 源文件行号（1-based）：用于结果汇总定位问题行
    data["__source_row__"] = [int(data_start_row_1based) + i for i in range(len(data))]
    unique_headers = list(unique_headers) + ["__source_row__"]

    # 只要指定了 max_rows 就加 __row_index__，避免键列重复（如 TIME 全空或相同）时多行被合并成 1 行
    if max_rows is not None and max_rows > 0:
        if len(data) > max_rows:
            data = data.iloc[:max_rows].reset_index(drop=True)
        if len(data) < max_rows:
            need = max_rows - len(data)
            empty = pd.DataFrame([[""] * len(unique_headers)] * need, columns=unique_headers)
            data = pd.concat([data, empty], ignore_index=True)
        data["__row_index__"] = [str(i) for i in range(len(data))]
        unique_headers = list(unique_headers) + ["__row_index__"]

    return data, unique_headers, header_rows


def _read_csv_tsv_safe(
    path: str,
    sep: Optional[str] = None,
    header_rows: Optional[int] = None,
    skip_top_rows: int = 0,
) -> Tuple[pd.DataFrame, List[str], int]:
    """
    读取 CSV/TSV。若 header_rows>1 则多行表头扁平化；否则单行表头。
    """
    ext = _normalize_ext(path)
    if sep is None:
        sep = "\t" if ext == ".tsv" else ","
    try:
        raw = pd.read_csv(path, sep=sep, header=None, dtype=str, encoding="utf-8")
    except UnicodeDecodeError:
        try:
            raw = pd.read_csv(path, sep=sep, header=None, dtype=str, encoding="gbk")
        except Exception as e:
            raise ParserError(f"读取 CSV/TSV 失败: {path}, 编码与格式错误: {e}") from e
    except Exception as e:
        raise ParserError(f"读取 CSV/TSV 失败: {path}, 错误: {e}") from e

    if raw.empty:
        return pd.DataFrame(), [], 0

    st = int(skip_top_rows or 0)
    if st > 0:
        if len(raw) <= st:
            raise ParserError(f"跳过顶部 {st} 行后无剩余数据，请减小「跳过顶部行数」")
        raw = raw.iloc[st:].reset_index(drop=True)

    if header_rows is None:
        header_rows = 1
    header_rows = min(header_rows, len(raw) - 1) if len(raw) > 1 else 1

    if header_rows == 1:
        headers = [str(c).strip() for c in raw.iloc[0].tolist()]
        data = raw.iloc[1:].reset_index(drop=True)
        data.columns = range(len(headers))
        data_start_row_1based = st + 2
    else:
        header_df = raw.iloc[: header_rows + 1]
        headers = _flatten_headers(header_df)
        data = raw.iloc[header_rows + 1 :].reset_index(drop=True)
        data.columns = range(len(headers))
        data_start_row_1based = st + (header_rows + 1) + 1

    headers = _simplify_flattened_headers(headers)

    seen = {}
    unique_headers = []
    for h in headers:
        h = str(h).strip()
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    data.columns = unique_headers
    data["__source_row__"] = [int(data_start_row_1based) + i for i in range(len(data))]
    unique_headers = list(unique_headers) + ["__source_row__"]
    return data, unique_headers, header_rows + 1


def _read_docx_safe(path: str) -> Tuple[pd.DataFrame, List[str], int]:
    """
    读取 Word 文档中的表格：取第一个表格，第一行作为表头，其余为数据。
    多表或复杂排版可后续扩展。
    """
    try:
        from docx import Document
    except ImportError:
        raise ParserError("请安装 python-docx: pip install python-docx") from None

    try:
        doc = Document(path)
    except Exception as e:
        raise ParserError(f"读取 Word 失败: {path}, 错误: {e}") from e

    tables = doc.tables
    if not tables:
        return pd.DataFrame(), [], 0

    table = tables[0]
    rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
    if not rows:
        return pd.DataFrame(), [], 0

    headers = [str(h) for h in rows[0]]
    data = pd.DataFrame(rows[1:], columns=range(len(headers)))
    seen = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    data.columns = unique_headers
    # Word：第 1 行为表头，数据从第 2 行开始
    data["__source_row__"] = [2 + i for i in range(len(data))]
    unique_headers = list(unique_headers) + ["__source_row__"]
    return data, unique_headers, 1


def load_table_from_file(
    path: str,
    *,
    header_rows: Optional[int] = None,
    max_rows: Optional[int] = None,
    skip_top_rows: int = 0,
    encoding: Optional[str] = None,
) -> Tuple[pd.DataFrame, List[str], int]:
    """
    根据文件扩展名自动选择解析器，加载表格数据。

    :param path: 文件路径
    :param header_rows: 表头占用行数（从跳过后的第1行起连续几行合并为列名）；None 表示自动检测（仅 Excel）
    :param max_rows: 仅 Excel：强制读取的数据行数（如 120000），留空则按文件实际范围
    :param skip_top_rows: 跳过文件顶部若干行后再识别表头（如第1行为标题、第2行为表头时填1）
    :param encoding: 主要用于 CSV/TSV
    :return: (DataFrame, 列名列表, 表头行数)
    """
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        raise ParserError(f"文件不存在: {path}")

    ext = _normalize_ext(path)
    if ext in (".xlsx", ".xls"):
        return _read_excel_safe(path, header_rows=header_rows, max_rows=max_rows, skip_top_rows=skip_top_rows)
    if ext in (".csv", ".tsv"):
        return _read_csv_tsv_safe(path, header_rows=header_rows, skip_top_rows=skip_top_rows)
    if ext == ".docx":
        return _read_docx_safe(path)
    if ext == ".et":
        # .et 为 WPS 表格格式，可后续接入转换或专用库
        raise ParserError("暂不支持 .et 格式，请先另存为 .xlsx 后使用")
    raise ParserError(f"不支持的文件格式: {ext}")


def get_columns_from_file(
    path: str,
    header_rows: Optional[int] = None,
    skip_top_rows: int = 0,
) -> List[str]:
    """
    仅获取表头列名，不加载全部数据，用于配置“对比字段”等场景。
    """
    df, cols, _ = load_table_from_file(path, header_rows=header_rows, skip_top_rows=skip_top_rows)
    return cols
